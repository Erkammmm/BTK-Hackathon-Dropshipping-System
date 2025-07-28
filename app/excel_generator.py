import os
import pandas as pd
from datetime import datetime
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelGenerator:
    def __init__(self):
        self.output_dir = "reports"
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def create_book_analysis_report(self, search_results: Dict, best_offer: Dict, gemini_analysis: Dict) -> str:
        """Kitap analizi için Excel raporu oluştur"""
        
        # Dosya adı oluştur
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        book_title = best_offer.get('title', 'kitap').replace(' ', '_')[:30]
        filename = f"kitap_analizi_{book_title}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Excel workbook oluştur
        wb = Workbook()
        
        # Ana sayfa - Özet
        self.create_summary_sheet(wb, best_offer, gemini_analysis)
        
        # Fiyat karşılaştırma sayfası
        self.create_price_comparison_sheet(wb, search_results, best_offer)
        
        # Kar analizi sayfası
        self.create_profit_analysis_sheet(wb, best_offer, gemini_analysis)
        
        # Detaylı analiz sayfası
        self.create_detailed_analysis_sheet(wb, gemini_analysis)
        
        # Excel dosyasını kaydet
        wb.save(filepath)
        
        return filepath
    
    def create_summary_sheet(self, wb: Workbook, best_offer: Dict, gemini_analysis: Dict):
        """Özet sayfası oluştur"""
        ws = wb.active
        ws.title = "Özet"
        
        # Başlık
        ws['A1'] = "KİTAP ANALİZ RAPORU"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:H1')
        
        # Kitap bilgileri
        ws['A3'] = "Kitap Adı:"
        ws['B3'] = best_offer.get('title', '')
        ws['A4'] = "Platform:"
        ws['B4'] = best_offer.get('platform', '')
        ws['A5'] = "En Uygun Fiyat:"
        ws['B5'] = f"{best_offer.get('price', 0)} TL"
        ws['A6'] = "URL:"
        ws['B6'] = best_offer.get('url', '')
        
        # Stil uygula
        for row in range(3, 7):
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Satış uygunluğu
        ws['A8'] = "SATIŞ UYGUNLUĞU"
        ws['A8'].font = Font(bold=True, size=14)
        ws['A8'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ws.merge_cells('A8:H8')
        
        # Kar analizi özeti
        profit_analysis = gemini_analysis.get('profit_analysis', '')
        if 'Satış Uygunluğu:' in profit_analysis:
            lines = profit_analysis.split('\n')
            row = 10
            for line in lines:
                if any(keyword in line for keyword in ['Satış Uygunluğu:', 'Kar Analizi:', 'Önerilen Fiyat:', 'Risk Değerlendirmesi:']):
                    ws[f'A{row}'] = line.strip()
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def create_price_comparison_sheet(self, wb: Workbook, search_results: Dict, best_offer: Dict):
        """Fiyat karşılaştırma sayfası oluştur"""
        ws = wb.create_sheet("Fiyat Karşılaştırma")
        
        # Başlık
        ws['A1'] = "FİYAT KARŞILAŞTIRMA TABLOSU"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Tablo başlıkları
        headers = ['Sıra', 'Kitap Adı', 'Platform', 'Fiyat (TL)', 'URL', 'Durum']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Verileri ekle
        all_results = []
        for platform, results in search_results.items():
            if platform != 'best_offer' and isinstance(results, list):
                for result in results:
                    all_results.append(result)
        
        # Fiyata göre sırala
        all_results.sort(key=lambda x: x.get('price', 0))
        
        for row, result in enumerate(all_results, 4):
            ws.cell(row=row, column=1, value=row-3)  # Sıra
            ws.cell(row=row, column=2, value=result.get('title', ''))
            ws.cell(row=row, column=3, value=result.get('platform', ''))
            ws.cell(row=row, column=4, value=result.get('price', 0))
            ws.cell(row=row, column=5, value=result.get('url', ''))
            
            # En ucuz olanı işaretle
            if result.get('price', 0) == best_offer.get('price', 0):
                ws.cell(row=row, column=6, value="EN UCUZ")
                ws.cell(row=row, column=6).fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            else:
                ws.cell(row=row, column=6, value="")
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 15
    
    def create_profit_analysis_sheet(self, wb: Workbook, best_offer: Dict, gemini_analysis: Dict):
        """Kar analizi sayfası oluştur"""
        ws = wb.create_sheet("Kar Analizi")
        
        # Başlık
        ws['A1'] = "KAR ANALİZİ"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:C1')
        
        # Maliyet hesaplama
        best_price = best_offer.get('price', 0)
        commission_rate = 0.21
        shipping_cost = 70
        profit_margin = 100
        
        # Hesaplamalar
        total_cost = best_price + shipping_cost
        commission_amount = (best_price + profit_margin) * commission_rate
        suggested_selling_price = total_cost + commission_amount + profit_margin
        net_profit = suggested_selling_price - total_cost
        profit_percentage = (profit_margin / suggested_selling_price) * 100 if suggested_selling_price > 0 else 0
        
        # Maliyet tablosu
        costs = [
            ['Alış Fiyatı', f"{best_price} TL"],
            ['Kargo Maliyeti', f"{shipping_cost} TL"],
            ['Toplam Maliyet', f"{total_cost} TL"],
            ['Komisyon (%21)', f"{commission_amount:.2f} TL"],
            ['Kar Marjı', f"{profit_margin} TL"],
            ['Önerilen Satış Fiyatı', f"{suggested_selling_price:.2f} TL"],
            ['Net Kar', f"{net_profit:.2f} TL"],
            ['Kar Yüzdesi', f"%{profit_percentage:.1f}"]
        ]
        
        for row, (item, value) in enumerate(costs, 3):
            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Satış uygunluğu değerlendirmesi
        ws['A12'] = "SATIŞ UYGUNLUĞU DEĞERLENDİRMESİ"
        ws['A12'].font = Font(bold=True, size=14)
        ws['A12'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ws.merge_cells('A12:C12')
        
        # Rekabet analizi
        profit_analysis = gemini_analysis.get('profit_analysis', '')
        if 'Satış Uygunluğu:' in profit_analysis:
            lines = profit_analysis.split('\n')
            row = 14
            for line in lines:
                if any(keyword in line for keyword in ['Satış Uygunluğu:', 'Kar Analizi:', 'Rekabet Durumu:', 'Risk Değerlendirmesi:']):
                    ws.cell(row=row, column=1, value=line.strip())
                    ws.cell(row=row, column=1).font = Font(bold=True)
                    row += 1
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
    
    def create_detailed_analysis_sheet(self, wb: Workbook, gemini_analysis: Dict):
        """Detaylı analiz sayfası oluştur"""
        ws = wb.create_sheet("Detaylı Analiz")
        
        # Başlık
        ws['A1'] = "DETAYLI GEMINI ANALİZİ"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:B1')
        
        # Analiz bölümleri
        sections = [
            ('Kitap Analizi', gemini_analysis.get('analysis', '')),
            ('SEO İçeriği', gemini_analysis.get('seo_content', '')),
            ('Satış Önerileri', gemini_analysis.get('sales_recommendation', '')),
            ('Özet', gemini_analysis.get('best_offer_summary', '')),
            ('Kar Analizi', gemini_analysis.get('profit_analysis', ''))
        ]
        
        row = 3
        for title, content in sections:
            ws.cell(row=row, column=1, value=title)
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            # İçeriği paragraflar halinde böl
            paragraphs = content.split('\n')
            for para in paragraphs:
                if para.strip():
                    ws.cell(row=row, column=1, value=para.strip())
                    row += 1
            
            row += 2  # Bölümler arası boşluk
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 20 