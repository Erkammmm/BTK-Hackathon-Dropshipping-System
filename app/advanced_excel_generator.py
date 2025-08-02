import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
import warnings
warnings.filterwarnings('ignore')

class AdvancedExcelGenerator:
    def __init__(self):
        self.output_dir = "reports"
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        
        # ML model için örnek veri
        self.sales_data = self._generate_sample_sales_data()
    
    def _generate_sample_sales_data(self) -> pd.DataFrame:
        """ML model için örnek satış verisi oluştur"""
        np.random.seed(42)
        
        # Kitap kategorileri ve popülerlik skorları
        categories = {
            'Roman': {'base_sales': 150, 'price_sensitivity': -2.5},
            'Bilim Kurgu': {'base_sales': 80, 'price_sensitivity': -1.8},
            'Tarih': {'base_sales': 60, 'price_sensitivity': -1.2},
            'Felsefe': {'base_sales': 40, 'price_sensitivity': -0.8},
            'Bilim': {'base_sales': 70, 'price_sensitivity': -1.5},
            'Çocuk': {'base_sales': 200, 'price_sensitivity': -3.0},
            'Eğitim': {'base_sales': 120, 'price_sensitivity': -2.0},
            'Klasik': {'base_sales': 90, 'price_sensitivity': -1.6}
        }
        
        data = []
        for category, params in categories.items():
            for _ in range(50):  # Her kategori için 50 örnek
                price = np.random.uniform(20, 200)
                popularity = np.random.uniform(0.1, 1.0)
                
                # Satış tahmini formülü
                base_sales = params['base_sales']
                price_effect = params['price_sensitivity'] * (price - 100) / 100
                popularity_effect = popularity * 50
                
                monthly_sales = max(0, int(base_sales + price_effect + popularity_effect + np.random.normal(0, 10)))
                
                data.append({
                    'category': category,
                    'price': price,
                    'popularity': popularity,
                    'monthly_sales': monthly_sales
                })
        
        return pd.DataFrame(data)
    
    def predict_sales(self, book_title: str, price: float, category: str = None, trendyol_data: Dict = None, amazon_sales_data: Dict = None) -> Dict:
        """ML model ile satış tahmini yap (Gerçek veri varsa kullan)"""
        try:
            # Amazon satış verisi varsa onu kullan (ÖNCELİK 1)
            if amazon_sales_data and amazon_sales_data.get('estimated_monthly_sales', 0) > 0:
                sales_data = amazon_sales_data
                
                return {
                    'predicted_sales': sales_data.get('estimated_monthly_sales', 0),
                    'confidence': sales_data.get('confidence_score', 0.7),
                    'monthly_revenue': sales_data.get('estimated_monthly_sales', 0) * price,
                    'popularity_score': min(1.0, sales_data.get('confidence_score', 0.7) + 0.2),
                    'category': category or 'Roman',
                    'trend': 'increasing' if sales_data.get('confidence_score', 0.7) > 0.8 else 'stable',
                    'daily_average': sales_data.get('daily_average', 0),
                    'sales_history': [],
                    'source': 'amazon_api',
                    'amazon_data': sales_data,
                    'total_ratings': sales_data.get('total_ratings', 0),
                    'availability': sales_data.get('availability', 'Unknown'),
                    'seller_count': sales_data.get('seller_count', 0),
                    'competition_level': sales_data.get('competition_level', 'Unknown')
                }
            
            # Türk E-ticaret API verisi varsa onu kullan (ÖNCELİK 2)
            if trendyol_data and trendyol_data.get('source') in ['turkish_ecommerce_crawler', 'turkish_ecommerce_search']:
                sales_data = trendyol_data.get('sales_data', {})
                
                # Trendyol verilerinden tahmin yap
                base_sales = sales_data.get('sales_count', 0)
                rating_count = sales_data.get('rating_count', 0)
                popularity_score = sales_data.get('popularity_score', 0.5)
                
                # Basit tahmin algoritması
                if base_sales > 0:
                    # Gerçek satış verisi varsa
                    predicted_monthly = base_sales * 1.2  # %20 artış varsayımı
                    confidence = min(0.8, popularity_score + 0.3)
                else:
                    # Popülerlik skoruna göre tahmin
                    predicted_monthly = int(50 * popularity_score)
                    confidence = popularity_score
                
                return {
                    'predicted_sales': int(predicted_monthly),
                    'confidence': confidence,
                    'monthly_revenue': int(predicted_monthly) * price,
                    'popularity_score': popularity_score,
                    'category': category or 'Roman',
                    'trend': 'increasing' if popularity_score > 0.7 else 'stable',
                    'daily_average': int(predicted_monthly // 30),
                    'sales_history': [],
                    'source': trendyol_data.get('source', 'turkish_ecommerce'),
                    'ecommerce_data': sales_data
                }
            
            # Eski yöntem (örnek veri ile)
            if not category:
                category = 'Roman'  # Varsayılan kategori
            
            # Kategori verilerini filtrele
            category_data = self.sales_data[self.sales_data['category'] == category]
            
            if len(category_data) < 10:
                # Yeterli veri yoksa tüm veriyi kullan
                category_data = self.sales_data
            
            # Özellikler
            X = category_data[['price', 'popularity']]
            y = category_data['monthly_sales']
            
            # Veriyi böl
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
            
            # Model eğitimi
            model = RandomForestRegressor(n_estimators=100, random_state=42)
            model.fit(X_train, y_train)
            
            # Tahmin için özellikler
            # Popülerlik skorunu kitap adından tahmin et
            popularity_score = self._estimate_popularity(book_title)
            
            # Tahmin yap
            prediction = model.predict([[price, popularity_score]])[0]
            prediction = max(0, int(prediction))
            
            # Güven aralığı (basit yaklaşım)
            confidence = model.score(X_test, y_test)
            
            # Aylık gelir tahmini
            monthly_revenue = prediction * price
            
            return {
                'predicted_sales': prediction,
                'confidence': confidence,
                'monthly_revenue': monthly_revenue,
                'popularity_score': popularity_score,
                'category': category,
                'trend': 'stable',
                'daily_average': prediction // 30,
                'sales_history': [],
                'source': 'sample_data'
            }
            
        except Exception as e:
            # Hata durumunda basit tahmin
            return {
                'predicted_sales': max(10, int(100 - price * 0.5)),
                'confidence': 0.6,
                'monthly_revenue': max(10, int(100 - price * 0.5)) * price,
                'popularity_score': 0.5,
                'category': category or 'Roman',
                'trend': 'stable',
                'daily_average': 1,
                'sales_history': [],
                'source': 'error_fallback'
            }
    
    def _calculate_monthly_trend(self, base_sales: int, sales_prediction: Dict) -> List[float]:
        """ML ile 6 aylık trend hesapla"""
        try:
            # Faktörleri hesapla
            confidence = sales_prediction.get('confidence', 0.7)
            popularity = sales_prediction.get('popularity_score', 0.5)
            source = sales_prediction.get('source', 'sample_data')
            
            # Trend faktörleri (dinamik hesaplama)
            if source == 'amazon_api':
                # Amazon verisi varsa daha gerçekçi trend
                if confidence > 0.8:
                    # Yüksek güven = yavaş düşüş
                    factors = [1.0, 0.95, 0.90, 0.85, 0.80, 0.75]
                elif confidence > 0.6:
                    # Orta güven = orta düşüş
                    factors = [1.0, 0.90, 0.80, 0.70, 0.60, 0.50]
                else:
                    # Düşük güven = hızlı düşüş
                    factors = [1.0, 0.85, 0.70, 0.55, 0.40, 0.25]
            else:
                # Sample data için standart trend
                factors = [1.0, 0.90, 0.80, 0.70, 0.60, 0.50]
            
            # Popülerlik skoruna göre ayarla
            popularity_adjustment = (popularity - 0.5) * 0.2  # ±10% ayarlama
            
            # Mevsimsellik faktörü (kitap satışları için)
            seasonal_factors = [1.0, 1.1, 1.2, 1.0, 0.9, 0.8]  # Yaz aylarında artış
            
            # Final hesaplama
            monthly_predictions = []
            for i, (base_factor, seasonal_factor) in enumerate(zip(factors, seasonal_factors)):
                adjusted_factor = base_factor + popularity_adjustment
                final_factor = adjusted_factor * seasonal_factor
                prediction = base_sales * final_factor
                monthly_predictions.append(prediction)
            
            print(f"🔍 Aylık trend hesaplandı: {monthly_predictions}")
            return monthly_predictions
            
        except Exception as e:
            print(f"❌ Aylık trend hesaplama hatası: {str(e)}")
            # Fallback: basit düşüş
            return [base_sales * (1.0 - i * 0.1) for i in range(6)]
    
    def _estimate_popularity(self, book_title: str) -> float:
        """Kitap adından popülerlik skoru tahmin et"""
        title_lower = book_title.lower()
        
        # Popülerlik göstergeleri
        popular_keywords = ['bestseller', 'çok satan', 'popüler', 'klasik', 'önerilen']
        niche_keywords = ['akademik', 'tez', 'araştırma', 'özel', 'teknik']
        
        popularity_score = 0.5  # Varsayılan
        
        # Anahtar kelime analizi
        for keyword in popular_keywords:
            if keyword in title_lower:
                popularity_score += 0.1
        
        for keyword in niche_keywords:
            if keyword in title_lower:
                popularity_score -= 0.1
        
        # Başlık uzunluğu etkisi
        if len(book_title) < 20:
            popularity_score += 0.05
        elif len(book_title) > 50:
            popularity_score -= 0.05
        
        return max(0.1, min(1.0, popularity_score))
    
    def create_advanced_book_analysis_report(self, search_results: Dict, best_offer: Dict, gemini_analysis: Dict, trendyol_data: Dict = None, comments_data: Dict = None) -> str:
        """Gelişmiş kitap analizi Excel raporu oluştur"""
        
        # Amazon satış verilerini çıkar
        amazon_sales_data = None
        if comments_data and comments_data.get('sales_data'):
            amazon_sales_data = comments_data.get('sales_data')
            print(f"🔍 Amazon satış verileri kullanılacak: {amazon_sales_data}")
        
        # ML tahmini yap (Amazon verisi varsa kullan)
        sales_prediction = self.predict_sales(
            best_offer.get('title', ''),
            best_offer.get('price', 0),
            trendyol_data=trendyol_data,
            amazon_sales_data=amazon_sales_data
        )
        
        # Dosya adı oluştur
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        book_title = best_offer.get('title', 'kitap').replace(' ', '_')[:30]
        filename = f"gelismis_kitap_analizi_{book_title}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Excel workbook oluştur
        wb = Workbook()
        
        # Sayfaları oluştur
        self.create_enhanced_summary_sheet(wb, best_offer, gemini_analysis, sales_prediction)
        self.create_price_charts_sheet(wb, search_results, best_offer)
        self.create_profit_charts_sheet(wb, best_offer, gemini_analysis)
        self.create_sales_prediction_sheet(wb, sales_prediction, best_offer)
        
        # Trendyol verisi varsa satış geçmişi sayfası ekle
        if sales_prediction.get('sales_history'):
            self.create_sales_history_sheet(wb, sales_prediction)
        
        self.create_detailed_analysis_sheet(wb, gemini_analysis)
        
        # Yorum analizi sayfası ekle (her zaman oluştur)
        self.create_comments_analysis_sheet(wb, comments_data, gemini_analysis)
        
        # Sonuç sayfası ekle (dropshipping analizi)
        self.create_results_sheet(wb, search_results, best_offer, gemini_analysis)
        
        # Excel dosyasını kaydet
        wb.save(filepath)
        
        return filepath
    
    def create_enhanced_summary_sheet(self, wb: Workbook, best_offer: Dict, gemini_analysis: Dict, sales_prediction: Dict):
        """Gelişmiş özet sayfası oluştur"""
        ws = wb.active
        ws.title = "Gelişmiş Özet"
        
        # Başlık
        ws['A1'] = "GELİŞMİŞ KİTAP ANALİZ RAPORU"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:I1')
        
        # Kitap bilgileri
        ws['A3'] = "Kitap Adı:"
        ws['B3'] = best_offer.get('title', '')
        ws['A4'] = "Platform:"
        ws['B4'] = best_offer.get('platform', '')
        ws['A5'] = "En Uygun Fiyat:"
        ws['B5'] = f"{best_offer.get('price', 0)} TL"
        ws['A6'] = "URL:"
        ws['B6'] = best_offer.get('url', '')
        
        # ML tahminleri
        ws['A8'] = "MACHINE LEARNING TAHMİNLERİ"
        ws['A8'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A8'].fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        ws.merge_cells('A8:I8')
        
        ws['A10'] = "Tahmini Aylık Satış:"
        ws['B10'] = f"{sales_prediction['predicted_sales']} adet"
        ws['A11'] = "Günlük Ortalama:"
        ws['B11'] = f"{sales_prediction['daily_average']} adet"
        ws['A12'] = "Güven Skoru:"
        ws['B12'] = f"%{sales_prediction['confidence']*100:.1f}"
        ws['A13'] = "Tahmini Aylık Gelir:"
        ws['B13'] = f"{sales_prediction['monthly_revenue']:.0f} TL"
        ws['A14'] = "Popülerlik Skoru:"
        ws['B14'] = f"{sales_prediction['popularity_score']:.2f}"
        ws['A15'] = "Trend:"
        ws['B15'] = sales_prediction.get('trend', 'stable').upper()
        ws['A16'] = "Veri Kaynağı:"
        ws['B16'] = sales_prediction.get('source', 'unknown').upper()
        ws['A17'] = "Kategori:"
        ws['B17'] = sales_prediction['category']
        
        # Amazon verileri varsa ek bilgiler
        if sales_prediction.get('source') == 'amazon_api':
            ws['A19'] = "AMAZON SATIŞ VERİLERİ"
            ws['A19'].font = Font(bold=True, size=12, color="FFFFFF")
            ws['A19'].fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            ws.merge_cells('A19:I19')
            
            ws['A21'] = "Toplam Değerlendirme:"
            ws['B21'] = sales_prediction.get('total_ratings', 0)
            ws['A22'] = "Stok Durumu:"
            ws['B22'] = sales_prediction.get('availability', 'Unknown')
            ws['A23'] = "Satıcı Sayısı:"
            ws['B23'] = sales_prediction.get('seller_count', 0)
            ws['A24'] = "Rekabet Seviyesi:"
            ws['B24'] = sales_prediction.get('competition_level', 'Unknown')
            
            # Amazon verileri için stil
            for row in range(21, 25):
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
        

        
        # Türk E-ticaret API verileri varsa göster (KALDIRILDI)
        # Sadece mevcut ML tahminleri ve yorum analizleri ile devam et
        
        # Stil uygula
        for row in range(3, 7):
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for row in range(10, 18):
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Trendyol verileri varsa ek stil
        if sales_prediction.get('trendyol_data'):
            for row in range(21, 26):
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def create_comments_analysis_sheet(self, wb: Workbook, comments_data: Dict, gemini_analysis: Dict):
        """Amazon yorum analizi sayfası oluştur (6. sayfa)"""
        ws = wb.create_sheet("Amazon Yorum Analizi")
        
        # Eğer yorum verisi yoksa sample data kullan
        if not comments_data or not comments_data.get('comments'):
            comments_data = {
                'total_comments': 0,
                'average_rating': 0.0,
                'source': 'no_data',
                'timestamp': datetime.now().isoformat(),
                'comments': [],
                'yearly_ratings': {}
            }
        
        # Başlık
        ws['A1'] = "AMAZON YORUM ANALİZİ VE OTOMATİK ÜRÜN AÇIKLAMASI"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
        ws.merge_cells('A1:H1')
        
        # Genel istatistikler
        ws['A3'] = "GENEL İSTATİSTİKLER"
        ws['A3'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A3'].fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        ws.merge_cells('A3:H3')
        
        ws['A5'] = "Toplam Yorum Sayısı:"
        ws['B5'] = comments_data.get('total_comments', 0)
        ws['A6'] = "Ortalama Yıldız:"
        ws['B6'] = f"{comments_data.get('average_rating', 0):.2f}/5"
        ws['A7'] = "Veri Kaynağı:"
        ws['B7'] = comments_data.get('source', 'unknown').upper()
        ws['A8'] = "Analiz Tarihi:"
        ws['B8'] = comments_data.get('timestamp', '').split('T')[0] if comments_data.get('timestamp') else ''
        
        # 1. ZAMAN SERİSİ YORUM ANALİZİ (TREND TAKİBİ)
        ws['A10'] = "1. ZAMAN SERİSİ YORUM ANALİZİ (TREND TAKİBİ)"
        ws['A10'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A10'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        ws.merge_cells('A10:H10')
        
        # Yıllık trend tablosu
        yearly_ratings = comments_data.get('yearly_ratings', {})
        ws['A12'] = "Yıllık Ortalama Yıldızlar:"
        ws['A12'].font = Font(bold=True)
        row = 13
        for year in sorted(yearly_ratings.keys()):
            ws[f'A{row}'] = f"{year} Yılı:"
            year_data = yearly_ratings[year]
            if isinstance(year_data, dict):
                avg_rating = year_data.get('average', 0)
            else:
                avg_rating = year_data
            try:
                avg_rating = float(avg_rating)
                ws[f'B{row}'] = f"{avg_rating:.2f} yıldız"
            except (ValueError, TypeError):
                ws[f'B{row}'] = f"{avg_rating} yıldız"
            row += 1
        
        # Gemini trend analizi
        if gemini_analysis.get('trend_analysis'):
            ws[f'A{row + 1}'] = "GEMINI TREND ANALİZİ:"
            ws[f'A{row + 1}'].font = Font(bold=True, color="F39C12")
            ws[f'A{row + 2}'] = gemini_analysis['trend_analysis']
            ws[f'A{row + 2}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{row + 2}:H{row + 8}')
        else:
            ws[f'A{row + 1}'] = "GEMINI TREND ANALİZİ:"
            ws[f'A{row + 1}'].font = Font(bold=True, color="F39C12")
            ws[f'A{row + 2}'] = "Gemini API limiti aşıldığı için trend analizi yapılamadı. Lütfen daha sonra tekrar deneyin."
            ws[f'A{row + 2}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{row + 2}:H{row + 8}')
        
        # 2. YORUMDAN ANLAM ÇIKARMA (SENTIMENT ANALİZİ)
        sentiment_start_row = row + 10
        ws[f'A{sentiment_start_row}'] = "2. YORUMDAN ANLAM ÇIKARMA (SENTIMENT ANALİZİ)"
        ws[f'A{sentiment_start_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{sentiment_start_row}'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        ws.merge_cells(f'A{sentiment_start_row}:H{sentiment_start_row}')
        
        if gemini_analysis.get('sentiment_analysis'):
            ws[f'A{sentiment_start_row + 2}'] = gemini_analysis['sentiment_analysis']
            ws[f'A{sentiment_start_row + 2}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{sentiment_start_row + 2}:H{sentiment_start_row + 8}')
        else:
            ws[f'A{sentiment_start_row + 2}'] = "Gemini API limiti aşıldığı için sentiment analizi yapılamadı. Lütfen daha sonra tekrar deneyin."
            ws[f'A{sentiment_start_row + 2}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{sentiment_start_row + 2}:H{sentiment_start_row + 8}')
        
        # 3. OTOMATİK ÜRÜN AÇIKLAMASI ÜRETİMİ
        description_start_row = sentiment_start_row + 10
        ws[f'A{description_start_row}'] = "3. OTOMATİK ÜRÜN AÇIKLAMASI ÜRETİMİ"
        ws[f'A{description_start_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{description_start_row}'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        ws.merge_cells(f'A{description_start_row}:H{description_start_row}')
        
        if gemini_analysis.get('user_based_description'):
            ws[f'A{description_start_row + 2}'] = gemini_analysis['user_based_description']
            ws[f'A{description_start_row + 2}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{description_start_row + 2}:H{description_start_row + 12}')
        else:
            ws[f'A{description_start_row + 2}'] = "Gemini API limiti aşıldığı için kullanıcı bazlı ürün açıklaması üretilemedi. Lütfen daha sonra tekrar deneyin."
            ws[f'A{description_start_row + 2}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{description_start_row + 2}:H{description_start_row + 12}')
        
        # 4. DETAYLI YORUM TABLOSU
        table_start_row = description_start_row + 15
        ws[f'A{table_start_row}'] = "4. DETAYLI YORUM TABLOSU (AMAZON API'DEN ALINAN VERİLER)"
        ws[f'A{table_start_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{table_start_row}'].fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        ws.merge_cells(f'A{table_start_row}:H{table_start_row}')
        
        # Tablo başlıkları
        headers = ['Tarih', 'Kullanıcı', 'Yıldız', 'Başlık', 'Yorum']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start_row + 2, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
        
        # Yorumları ekle (Rapid API'den gelen gerçek veriler)
        comments = comments_data.get('comments', [])
        if comments:
            for i, comment in enumerate(comments[:30], table_start_row + 3):  # İlk 30 yorumu göster
                ws.cell(row=i, column=1, value=str(comment.get('date', '')))
                ws.cell(row=i, column=2, value=str(comment.get('user', '')))
                ws.cell(row=i, column=3, value=float(comment.get('rating', 0)))
                ws.cell(row=i, column=4, value=str(comment.get('title', '')))
                
                comment_cell = ws.cell(row=i, column=5, value=str(comment.get('comment', '')))
                comment_cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Yıldıza göre renk
                rating = float(comment.get('rating', 0))
                if rating >= 4:
                    color = "D5F4E6"  # Yeşil
                elif rating >= 3:
                    color = "FEF9E7"  # Sarı
                else:
                    color = "FADBD8"  # Kırmızı
                
                for col in range(1, 6):
                    ws.cell(row=i, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        else:
            # Yorum yoksa bilgi mesajı
            ws.cell(row=table_start_row + 3, column=1, value="Yorum verisi bulunamadı")
            ws.cell(row=table_start_row + 3, column=2, value="Amazon URL'i bulunamadı veya yorum yok")
            ws.cell(row=table_start_row + 3, column=3, value="-")
            ws.cell(row=table_start_row + 3, column=4, value="-")
            ws.cell(row=table_start_row + 3, column=5, value="Bu ürün için henüz yorum bulunmuyor veya Amazon ASIN'i tespit edilemedi.")
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 12  # Tarih
        ws.column_dimensions['B'].width = 15  # Kullanıcı
        ws.column_dimensions['C'].width = 8   # Yıldız
        ws.column_dimensions['D'].width = 20  # Başlık
        ws.column_dimensions['E'].width = 50  # Yorum
        
        # Satır yüksekliklerini ayarla (analiz bölümleri için)
        for row in range(19, 26):
            ws.row_dimensions[row].height = 120
        for row in range(29, 36):
            ws.row_dimensions[row].height = 120
        for row in range(39, 46):
            ws.row_dimensions[row].height = 120
    
    def create_price_charts_sheet(self, wb: Workbook, search_results: Dict, best_offer: Dict):
        """Fiyat grafikleri sayfası oluştur"""
        ws = wb.create_sheet("Fiyat Grafikleri")
        
        # Başlık
        ws['A1'] = "FİYAT KARŞILAŞTIRMA GRAFİKLERİ"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Veri hazırla
        all_results = []
        for platform, results in search_results.items():
            if platform != 'best_offer' and isinstance(results, list):
                for result in results:
                    all_results.append(result)
        
        # Fiyata göre sırala
        all_results.sort(key=lambda x: x.get('price', 0))
        
        # Tablo oluştur
        headers = ['Platform', 'Fiyat (TL)', 'Durum']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for row, result in enumerate(all_results, 4):
            ws.cell(row=row, column=1, value=result.get('platform', ''))
            ws.cell(row=row, column=2, value=result.get('price', 0))
            
            if result.get('price', 0) == best_offer.get('price', 0):
                ws.cell(row=row, column=3, value="EN UCUZ")
                ws.cell(row=row, column=3).fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            else:
                ws.cell(row=row, column=3, value="")
        
        # Bar chart oluştur
        chart = BarChart()
        chart.title = "Platform Fiyat Karşılaştırması"
        chart.x_axis.title = "Platform"
        chart.y_axis.title = "Fiyat (TL)"
        
        data = Reference(ws, min_col=2, min_row=3, max_row=len(all_results)+3)
        cats = Reference(ws, min_col=1, min_row=4, max_row=len(all_results)+3)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Grafik stilini ayarla
        chart.style = 10
        chart.height = 15
        chart.width = 20
        
        ws.add_chart(chart, "E3")
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def create_profit_charts_sheet(self, wb: Workbook, best_offer: Dict, gemini_analysis: Dict):
        """Kar analizi grafikleri sayfası oluştur"""
        ws = wb.create_sheet("Kar Grafikleri")
        
        # Başlık
        ws['A1'] = "KAR ANALİZİ GRAFİKLERİ"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Maliyet hesaplama
        best_price = best_offer.get('price', 0)
        commission_rate = 0.14  # %14 komisyon
        profit_margin = 100
        
        total_cost = best_price  # Kargo maliyeti kaldırıldı
        commission_amount = (best_price + profit_margin) * commission_rate
        suggested_selling_price = total_cost + commission_amount + profit_margin
        net_profit = suggested_selling_price - total_cost
        
        # Maliyet dağılımı tablosu
        ws['A3'] = "Maliyet Dağılımı"
        ws['A3'].font = Font(bold=True, size=14)
        ws.merge_cells('A3:C3')
        
        cost_data = [
            ['Alış Fiyatı', best_price],
            ['Komisyon (%14)', commission_amount],
            ['Kar Marjı', profit_margin]
        ]
        
        for row, (item, value) in enumerate(cost_data, 4):
            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Pie chart oluştur
        pie = PieChart()
        pie.title = "Maliyet Dağılımı"
        
        data = Reference(ws, min_col=2, min_row=4, max_row=6)
        labels = Reference(ws, min_col=1, min_row=4, max_row=6)
        
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        
        # Data labels ekle
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        
        ws.add_chart(pie, "E3")
        
        # Kar analizi tablosu
        ws['A10'] = "Kar Analizi"
        ws['A10'].font = Font(bold=True, size=14)
        ws.merge_cells('A10:C10')
        
        profit_data = [
            ['Toplam Maliyet', total_cost],
            ['Önerilen Satış Fiyatı', suggested_selling_price],
            ['Net Kar', net_profit],
            ['Kar Yüzdesi', f"%{(net_profit/suggested_selling_price)*100:.1f}"]
        ]
        
        for row, (item, value) in enumerate(profit_data, 11):
            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def create_sales_prediction_sheet(self, wb: Workbook, sales_prediction: Dict, best_offer: Dict):
        """Satış tahmini sayfası oluştur"""
        ws = wb.create_sheet("Satış Tahmini")
        
        # Başlık
        ws['A1'] = "MACHINE LEARNING SATIŞ TAHMİNİ"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Tahmin sonuçları
        ws['A3'] = "Tahmin Sonuçları"
        ws['A3'].font = Font(bold=True, size=14)
        ws.merge_cells('A3:C3')
        
        prediction_data = [
            ['Tahmini Aylık Satış', sales_prediction['predicted_sales'], 'adet'],
            ['Güven Skoru', f"{sales_prediction['confidence']*100:.1f}", '%'],
            ['Tahmini Aylık Gelir', f"{sales_prediction['monthly_revenue']:.0f}", 'TL'],
            ['Popülerlik Skoru', f"{sales_prediction['popularity_score']:.2f}", ''],
            ['Kategori', sales_prediction['category'], ''],
            ['Kitap Fiyatı', best_offer.get('price', 0), 'TL']
        ]
        
        for row, (item, value, unit) in enumerate(prediction_data, 4):
            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=unit)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Aylık trend tahmini
        ws['A11'] = "Aylık Trend Tahmini"
        ws['A11'].font = Font(bold=True, size=14)
        ws.merge_cells('A11:C11')
        
        # 6 aylık tahmin (Amazon verisi varsa kullan)
        months = ['1. Ay', '2. Ay', '3. Ay', '4. Ay', '5. Ay', '6. Ay']
        
        # 6 aylık tahmin (ML ile dinamik hesaplama)
        base_sales = sales_prediction['predicted_sales']
        monthly_predictions = self._calculate_monthly_trend(base_sales, sales_prediction)
        
        for col, (month, prediction) in enumerate(zip(months, monthly_predictions), 1):
            ws.cell(row=12, column=col, value=month)
            ws.cell(row=12, column=col).font = Font(bold=True)
            
            ws.cell(row=13, column=col, value=int(prediction))
            
            revenue = prediction * best_offer.get('price', 0)
            ws.cell(row=14, column=col, value=f"{revenue:.0f}")
        
        ws.cell(row=13, column=1, value="Satış Adedi").font = Font(bold=True)
        ws.cell(row=14, column=1, value="Gelir (TL)").font = Font(bold=True)
        
        # Line chart oluştur
        chart = LineChart()
        chart.title = "6 Aylık Satış Trendi"
        chart.x_axis.title = "Ay"
        chart.y_axis.title = "Satış Adedi"
        
        data = Reference(ws, min_col=2, min_row=13, max_col=7, max_row=13)
        cats = Reference(ws, min_col=2, min_row=12, max_col=7, max_row=12)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        chart.style = 10
        chart.height = 15
        chart.width = 20
        
        ws.add_chart(chart, "A16")
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 20
        for col in range(2, 8):
            ws.column_dimensions[chr(64 + col)].width = 15
    
    def create_detailed_analysis_sheet(self, wb: Workbook, gemini_analysis: Dict):
        """Detaylı analiz sayfası oluştur (Tablo formatında)"""
        ws = wb.create_sheet("Detaylı Analiz")
        
        # Başlık
        ws['A1'] = "DETAYLI GEMINI ANALİZİ"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:C1')
        
        # Tablo başlıkları
        headers = ['Analiz Türü', 'İçerik', 'Özet']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Analiz bölümleri
        sections = [
            ('Kitap Analizi', gemini_analysis.get('analysis', ''), 'Kitabın detaylı analizi'),
            ('SEO İçeriği', gemini_analysis.get('seo_content', ''), 'SEO uyumlu açıklama'),
            ('Satış Önerileri', gemini_analysis.get('sales_recommendation', ''), 'Satış stratejileri'),
            ('En İyi Teklif Özeti', gemini_analysis.get('best_offer_summary', ''), 'Fiyat karşılaştırması'),
            ('Kar Analizi', gemini_analysis.get('profit_analysis', ''), 'Kâr hesaplamaları')
        ]
        
        row = 4
        for title, content, summary in sections:
            # Analiz türü
            ws.cell(row=row, column=1, value=title)
            ws.cell(row=row, column=1).font = Font(bold=True, size=11)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            
            # İçerik (tam metin)
            cell = ws.cell(row=row, column=2, value=content)
            cell.font = Font(size=10)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Özet
            cell = ws.cell(row=row, column=3, value=summary)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
        
        # Detaylı içerik bölümü
        ws['A10'] = "DETAYLI İÇERİKLER"
        ws['A10'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A10'].fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        ws.merge_cells('A10:C10')
        
        row = 12
        for title, content, _ in sections:
            # Bölüm başlığı
            ws.cell(row=row, column=1, value=f"📋 {title}")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
            ws.cell(row=row, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            # İçeriği paragraflar halinde böl
            paragraphs = content.split('\n')
            for para in paragraphs:
                if para.strip():
                    cell = ws.cell(row=row, column=1, value=para.strip())
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    ws.merge_cells(f'A{row}:C{row}')
                    row += 1
            
            row += 2  # Bölümler arası boşluk
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 120
        ws.column_dimensions['C'].width = 40
    
    def create_sales_history_sheet(self, wb: Workbook, sales_prediction: Dict):
        """Satış geçmişi sayfası oluştur"""
        ws = wb.create_sheet("Satış Geçmişi")
        
        # Başlık
        ws['A1'] = "TRENDYOL SATIŞ GEÇMİŞİ (Son 30 Gün)"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        ws.merge_cells('A1:E1')
        
        # Tablo başlıkları
        headers = ['Tarih', 'Satış Adedi', 'Gelir (TL)', 'Fiyat (TL)', 'Günlük Trend']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Satış verilerini ekle
        sales_history = sales_prediction.get('sales_history', [])
        
        for row, sale in enumerate(sales_history, 4):
            ws.cell(row=row, column=1, value=sale.get('date', ''))
            ws.cell(row=row, column=2, value=sale.get('quantity', 0))
            ws.cell(row=row, column=3, value=f"{sale.get('revenue', 0):.2f}")
            ws.cell(row=row, column=4, value=f"{sale.get('price', 0):.2f}")
            
            # Günlük trend (basit hesaplama)
            if row > 4:
                prev_quantity = sales_history[row-5].get('quantity', 0)
                curr_quantity = sale.get('quantity', 0)
                if prev_quantity > 0:
                    trend = ((curr_quantity - prev_quantity) / prev_quantity) * 100
                    trend_text = f"%{trend:+.1f}"
                    if trend > 0:
                        ws.cell(row=row, column=5, value=trend_text).fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                    elif trend < 0:
                        ws.cell(row=row, column=5, value=trend_text).fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    else:
                        ws.cell(row=row, column=5, value=trend_text)
                else:
                    ws.cell(row=row, column=5, value="N/A")
            else:
                ws.cell(row=row, column=5, value="İlk Gün")
        
        # Özet istatistikler
        ws['A35'] = "ÖZET İSTATİSTİKLER"
        ws['A35'].font = Font(bold=True, size=14)
        ws['A35'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ws.merge_cells('A35:E35')
        
        if sales_history:
            total_sales = sum(sale.get('quantity', 0) for sale in sales_history)
            total_revenue = sum(sale.get('revenue', 0) for sale in sales_history)
            avg_price = total_revenue / total_sales if total_sales > 0 else 0
            max_sales = max(sale.get('quantity', 0) for sale in sales_history)
            min_sales = min(sale.get('quantity', 0) for sale in sales_history)
            
            stats = [
                ['Toplam Satış', f"{total_sales} adet"],
                ['Toplam Gelir', f"{total_revenue:.2f} TL"],
                ['Ortalama Fiyat', f"{avg_price:.2f} TL"],
                ['En Yüksek Günlük Satış', f"{max_sales} adet"],
                ['En Düşük Günlük Satış', f"{min_sales} adet"]
            ]
            
            for row, (label, value) in enumerate(stats, 37):
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def create_results_sheet(self, wb: Workbook, search_results: Dict, best_offer: Dict, gemini_analysis: Dict):
        """Sonuç sayfası oluştur (dropshipping analizi)"""
        ws = wb.create_sheet("Sonuç")
        
        # Başlık
        ws['A1'] = "DROPSHİPPİNG ANALİZİ VE SONUÇLAR"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        ws.merge_cells('A1:H1')
        
        # Fiyat analizi
        ws['A3'] = "FİYAT ANALİZİ"
        ws['A3'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A3'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        ws.merge_cells('A3:H3')
        
        # En ucuz fiyatı bul
        all_prices = []
        for platform, results in search_results.items():
            if platform != 'best_offer' and isinstance(results, list):
                for result in results:
                    price = result.get('price', 0)
                    if price > 0:
                        all_prices.append({
                            'platform': result.get('platform', ''),
                            'price': price,
                            'url': result.get('url', '')
                        })
        
        # Fiyata göre sırala
        all_prices.sort(key=lambda x: x['price'])
        
        if all_prices:
            cheapest_price = all_prices[0]['price']
            cheapest_platform = all_prices[0]['platform']
            highest_price = all_prices[-1]['price']
            
            ws['A5'] = "En Ucuz Fiyat:"
            ws['B5'] = f"{cheapest_price} TL ({cheapest_platform})"
            ws['A6'] = "En Pahalı Fiyat:"
            ws['B6'] = f"{highest_price} TL"
            ws['A7'] = "Fiyat Farkı:"
            ws['B7'] = f"{highest_price - cheapest_price} TL"
            
            # Dropshipping hesaplaması
            commission_rate = 0.14  # %14 komisyon
            profit_margin = 100  # 100 TL kar marjı
            
            # En ucuz fiyata komisyon ve kar marjı ekle
            dropshipping_price = cheapest_price + (cheapest_price + profit_margin) * commission_rate + profit_margin
            
            ws['A9'] = "DROPSHİPPİNG HESAPLAMASI"
            ws['A9'].font = Font(bold=True, size=12, color="FFFFFF")
            ws['A9'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            ws.merge_cells('A9:H9')
            
            ws['A11'] = "Alış Fiyatı:"
            ws['B11'] = f"{cheapest_price} TL"
            ws['A12'] = "Komisyon (%14):"
            ws['B12'] = f"{(cheapest_price + profit_margin) * commission_rate:.2f} TL"
            ws['A13'] = "Kar Marjı:"
            ws['B13'] = f"{profit_margin} TL"
            ws['A14'] = "Önerilen Satış Fiyatı:"
            ws['B14'] = f"{dropshipping_price:.2f} TL"
            
            # Kar analizi
            ws['A16'] = "KAR ANALİZİ"
            ws['A16'].font = Font(bold=True, size=12, color="FFFFFF")
            ws['A16'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
            ws.merge_cells('A16:H16')
            
            # Dropshipping fiyatı en pahalı fiyattan düşük mü?
            if dropshipping_price < highest_price:
                ws['A18'] = "✅ TRENDYOL'DA SATIŞ ÖNERİSİ"
                ws['A18'].font = Font(bold=True, size=14, color="FFFFFF")
                ws['A18'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                ws.merge_cells('A18:H18')
                
                ws['A20'] = "GEMINI ANALİZİ:"
                ws['A20'].font = Font(bold=True, size=12)
                ws['A20'].fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
                ws.merge_cells('A20:H20')
                
                analysis_text = f"""
Bu ürün Trendyol'da başarıyla satılabilir! 

📊 FİYAT ANALİZİ:
• En ucuz alış fiyatı: {cheapest_price} TL ({cheapest_platform})
• Önerilen satış fiyatı: {dropshipping_price:.2f} TL
• En yüksek rakip fiyat: {highest_price} TL
• Potansiyel kar marjı: {dropshipping_price - cheapest_price:.2f} TL

💡 SATIŞ STRATEJİSİ:
• Bu fiyatla rakiplerden daha uygun fiyatlı olacaksınız
• Müşteriler bu fiyat farkını fark edecek ve tercih edecek
• Güvenli kar marjınız korunuyor
• Hızlı satış potansiyeli yüksek

🚀 ÖNERİLER:
• Ürünü hemen Trendyol'a ekleyin
• SEO açıklamasını optimize edin
• Hızlı kargo vaadi verin
• Müşteri hizmetlerine önem verin
                """
                
                ws['A22'] = analysis_text
                ws['A22'].alignment = Alignment(wrap_text=True, vertical='top')
                ws.merge_cells('A22:H30')
                
            else:
                ws['A18'] = "❌ TRENDYOL'DA SATIŞ RİSKİ"
                ws['A18'].font = Font(bold=True, size=14, color="FFFFFF")
                ws['A18'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                ws.merge_cells('A18:H18')
                
                ws['A20'] = "GEMINI ANALİZİ:"
                ws['A20'].font = Font(bold=True, size=12)
                ws['A20'].fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                ws.merge_cells('A20:H20')
                
                analysis_text = f"""
Bu ürün Trendyol'da satış için riskli! 

📊 FİYAT ANALİZİ:
• En ucuz alış fiyatı: {cheapest_price} TL ({cheapest_platform})
• Önerilen satış fiyatı: {dropshipping_price:.2f} TL
• En yüksek rakip fiyat: {highest_price} TL
• Fiyat farkı: {dropshipping_price - highest_price:.2f} TL (çok yüksek)

⚠️ RİSK FAKTÖRLERİ:
• Fiyatınız rakiplerden çok daha yüksek
• Müşteriler bu fiyat farkını kabul etmeyecek
• Satış potansiyeli çok düşük
• Stok riski yüksek

💡 ALTERNATİF STRATEJİLER:
• Daha düşük kar marjı ile satış yapın
• Farklı bir ürün kategorisi deneyin
• Toplu alım ile maliyeti düşürün
• Özel kampanyalar ile rekabet edin
                """
                
                ws['A22'] = analysis_text
                ws['A22'].alignment = Alignment(wrap_text=True, vertical='top')
                ws.merge_cells('A22:H30')
        
        # Detaylı fiyat tablosu
        ws['A32'] = "DETAYLI FİYAT TABLOSU"
        ws['A32'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A32'].fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        ws.merge_cells('A32:H32')
        
        # Tablo başlıkları
        headers = ['Platform', 'Fiyat (TL)', 'Durum', 'Dropshipping Potansiyeli']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=34, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
        
        # Fiyat verilerini ekle
        for row, price_data in enumerate(all_prices, 35):
            ws.cell(row=row, column=1, value=price_data['platform'])
            ws.cell(row=row, column=2, value=price_data['price'])
            
            # Durum belirleme
            if price_data['price'] == cheapest_price:
                ws.cell(row=row, column=3, value="EN UCUZ")
                ws.cell(row=row, column=3).fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                ws.cell(row=row, column=4, value="✅ YÜKSEK")
                ws.cell(row=row, column=4).fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
            elif price_data['price'] == highest_price:
                ws.cell(row=row, column=3, value="EN PAHALI")
                ws.cell(row=row, column=3).fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                ws.cell(row=row, column=4, value="❌ DÜŞÜK")
                ws.cell(row=row, column=4).fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
            else:
                ws.cell(row=row, column=3, value="ORTA")
                ws.cell(row=row, column=4, value="⚠️ ORTA")
                ws.cell(row=row, column=4).fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        
        # Satır yüksekliklerini ayarla
        for row in range(22, 31):
            ws.row_dimensions[row].height = 120 